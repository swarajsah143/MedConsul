#!/usr/bin/env python3
"""
extract_fees_from_sites.py — MedCounsel fee-gap filler.

Reads the "data gaps" workbook, and for every college that has a website link
(optionally only those still MISSING fees), it:
  1. fetches the homepage,
  2. discovers fee / prospectus / admission / hostel subpages and PDFs,
  3. extracts their text (HTML + PDF),
  4. asks a Groq LLM (OpenAI-compatible) to pull structured fee figures,
  5. checkpoints each result to a JSONL (so the run is resumable),
  6. writes a clean output workbook: university name + fee breakdown.

Env: reads AI_API_KEY / AI_API_BASE_URL / AI_MODEL from the repo-root .env.
Run inside the scratchpad venv that has: requests beautifulsoup4 lxml pypdf openai openpyxl
"""
import argparse, io, json, os, re, sys, threading, time
from collections import deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin, urlparse

import requests
import urllib3
from bs4 import BeautifulSoup
import openpyxl

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ----------------------------------------------------------------------------- env
def load_env():
    env = {}
    p = os.path.join(REPO, ".env")
    if os.path.exists(p):
        for line in open(p, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k] = v.strip().strip('"').strip("'")
    return env

ENV = load_env()
AI_KEY = ENV.get("AI_API_KEY") or os.environ.get("AI_API_KEY", "")
AI_BASE = ENV.get("AI_API_BASE_URL", "https://api.groq.com/openai/v1")
AI_MODEL = ENV.get("AI_MODEL", "llama-3.3-70b-versatile")

# --- separate extraction credentials -----------------------------------------
# So bulk extraction NEVER starves the live app: use a dedicated Groq key (from a
# SEPARATE account — Groq rate-limits per account, so a 2nd key on the same
# account shares the same budget). Precedence: env var > data/.env.extract file >
# fall back to the app key. The .env.extract file is git-ignored and survives the
# repo-root .env reverting.
def _load_extract_env():
    f = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env.extract")
    kv = {}
    if os.path.exists(f):
        for line in open(f, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                kv[k.strip()] = v.strip().strip('"').strip("'")
    return kv

_EXT = _load_extract_env()
def _ext(name, default=""):
    return (os.environ.get(f"EXTRACT_{name}")
            or _EXT.get(f"EXTRACT_{name}") or _EXT.get(name) or default)

# extraction key/base — falls back to the app's if no separate key is configured
EXTRACT_KEY = _ext("AI_API_KEY", AI_KEY)
EXTRACT_BASE = _ext("AI_API_BASE_URL", AI_BASE)
USING_SEPARATE_KEY = bool(EXTRACT_KEY and EXTRACT_KEY != AI_KEY)

# ----------------------------------------------------------------------------- fetch
UA = ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) "
      "Chrome/125.0 Safari/537.36")
HEADERS = {"User-Agent": UA, "Accept": "text/html,application/pdf,*/*"}

# strong fee signals (page is very likely the fee page)
FEE_STRONG = ["fee-structure", "feestructure", "fee structure", "fees", "fee", "tuition",
              "tution", "prospectus", "tariff", "hostel fee", "fee-details"]
# hub pages that usually link onward to the fee page
HUB_WORDS = ["admission", "academics", "student", "hostel", "accommodation", "download",
             "important-links", "quick-links", "mbbs", "bds"]
# pages we must NOT follow (they contain rupee amounts but are not fees)
NEG_WORDS = ["recruit", "vacancy", "career", "tender", "procure", "quotation", "e-tender",
             "corrigendum", "advertisement", "walk-in", "walkin", "notification", "result",
             "news", "press", "gallery", "syllabus", "faculty", "contact", "grievance",
             "rti", "circular", "notice", "event", "seminar", "conference", "purchase",
             "supply", "bid", "auction", "empanel", "nit"]
# common paths worth probing directly
PROBE_PATHS = ["fee-structure", "fees", "fee", "fee-structure/", "admission", "admissions",
               "prospectus", "fee-details", "student-corner", "academics"]
# amount patterns: ₹, Rs, INR, or Indian lakh grouping like 1,23,456 or 5+ digit runs
AMOUNT_RE = re.compile(r"(₹|\brs\.?\b|\binr\b|\d{1,2},\d{2},\d{3}|\b\d{5,}\b)", re.I)

def norm_url(u: str) -> str:
    u = (u or "").strip()
    if not u:
        return ""
    if "://" not in u:
        u = "http://" + u
    return u

def get(url, timeout=20, max_bytes=4_000_000):
    r = requests.get(url, headers=HEADERS, timeout=timeout, verify=False,
                     allow_redirects=True, stream=True)
    r.raise_for_status()
    ctype = r.headers.get("Content-Type", "").lower()
    buf = b""
    for chunk in r.iter_content(65536):
        buf += chunk
        if len(buf) > max_bytes:
            break
    return r.url, ctype, buf

def html_text(soup: BeautifulSoup) -> str:
    for t in soup(["script", "style", "noscript", "svg"]):
        t.decompose()
    return re.sub(r"\n{3,}", "\n\n", soup.get_text("\n", strip=True))

def pdf_text(data: bytes) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(data))
        out = []
        for page in reader.pages[:15]:
            out.append(page.extract_text() or "")
        return "\n".join(out)
    except Exception:
        return ""

def score_link(href, label):
    """Return (kind, score): kind in {'fee','hub',''}. Negative words veto the link."""
    s = (href + " " + label).lower()
    if any(n in s for n in NEG_WORDS):
        return "", -100
    fee = 0
    for k in FEE_STRONG:
        if k in s:
            fee += 6 if k in ("fee-structure", "feestructure", "fee structure",
                              "fees", "prospectus", "fee-details") else 4
    hub = sum(2 for k in HUB_WORDS if k in s)
    if href.lower().endswith(".pdf") and fee > 0:
        fee += 3          # a PDF explicitly about fees is gold
    if fee > 0:
        return "fee", fee + hub
    if hub > 0:
        return "hub", hub
    return "", 0

def _links(base_url, soup):
    """Yield (kind, score, url, is_pdf) for scored candidate links on a page."""
    seen = set()
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
            continue
        full = urljoin(base_url, href).split("#")[0]
        if urlparse(full).scheme not in ("http", "https"):
            continue
        if full in seen:
            continue
        seen.add(full)
        label = a.get_text(" ", strip=True)[:90]
        kind, sc = score_link(full, label)
        if sc <= 0:
            continue
        yield kind, sc, full, full.lower().endswith(".pdf")

def _fetch_text(url, timeout=18):
    fu, ct, bd = get(url, timeout=timeout)
    if "pdf" in ct or fu.lower().endswith(".pdf"):
        return fu, pdf_text(bd)
    try:
        s = BeautifulSoup(bd, "lxml")
    except Exception:
        s = BeautifulSoup(bd, "html.parser")
    return fu, html_text(s), s

def gather_pages(site, log, max_fetches=8):
    """Homepage + probed paths + fee links + 1 hop through hub pages -> (url,text) list."""
    pages, fetched = [], 0
    base = norm_url(site)
    root = f"{urlparse(base).scheme}://{urlparse(base).netloc}"
    try:
        final, ctype, body = get(base)
        fetched += 1
    except Exception as e:
        return pages, f"homepage_error: {type(e).__name__}: {str(e)[:110]}"
    if "pdf" in ctype:
        return [(final, pdf_text(body)[:8000])], "ok(pdf-home)"
    try:
        soup = BeautifulSoup(body, "lxml")
    except Exception:
        soup = BeautifulSoup(body, "html.parser")
    pages.append((final, html_text(soup)[:6000]))

    fee_links, hub_links, done = [], [], {final}
    for kind, sc, url, is_pdf in _links(final, soup):
        (fee_links if kind == "fee" else hub_links).append((sc, url, is_pdf))
    fee_links.sort(reverse=True); hub_links.sort(reverse=True)

    def grab(url):
        nonlocal fetched
        if url in done or fetched >= max_fetches:
            return None
        done.add(url)
        try:
            res = _fetch_text(url)
            fetched += 1
            fu, txt = res[0], res[1]
            if txt and txt.strip():
                pages.append((fu, txt[:7000]))
            return res
        except Exception:
            return None

    # 1) direct fee links from homepage (best signal)
    for _sc, url, _pdf in fee_links[:4]:
        grab(url)
    # 2) probe common fee paths
    for p in PROBE_PATHS:
        if fetched >= max_fetches:
            break
        grab(f"{root}/{p}")
    # 3) hop through up to 2 hub pages, harvesting fee links found there
    for _sc, hub, _pdf in hub_links[:2]:
        if fetched >= max_fetches:
            break
        res = grab(hub)
        if res and len(res) == 3:
            sub = sorted(((sc, u, ip) for k, sc, u, ip in _links(res[0], res[2])
                          if k == "fee"), reverse=True)
            for _s2, u2, _p2 in sub[:2]:
                if fetched >= max_fetches:
                    break
                grab(u2)
    return pages, "ok"

# ----------------------------------------------------------------------------- LLM
import requests as _rq

# Extraction model can differ from the chat model. 8b-instant has far higher
# throughput than 70b and is plenty for pulling a fee figure out of text.
EXTRACT_MODEL = _ext("AI_MODEL", os.environ.get("AI_EXTRACT_MODEL", AI_MODEL))

class RateLimiter:
    """Requests-per-minute limiter that also honours Groq's live token budget.

    After each call we read x-ratelimit-remaining-tokens / reset headers and, if
    the remaining token budget is low, park all callers until the bucket refills.
    This adapts to the real (free-tier) TPM instead of guessing.
    """
    def __init__(self, rpm):
        self.rpm = rpm
        self.calls = deque()
        self.lock = threading.Lock()
        self.sleep_until = 0.0
    def acquire(self):
        while True:
            with self.lock:
                now = time.time()
                if now < self.sleep_until:
                    wait = self.sleep_until - now
                else:
                    while self.calls and now - self.calls[0] > 60:
                        self.calls.popleft()
                    if len(self.calls) < self.rpm:
                        self.calls.append(now)
                        return
                    wait = 60 - (now - self.calls[0]) + 0.05
            time.sleep(max(wait, 0.1))
    def note_headers(self, headers):
        try:
            rem = int(headers.get("x-ratelimit-remaining-tokens", "999999"))
        except ValueError:
            return
        if rem < 2500:                       # getting close to the TPM wall
            reset = _parse_reset(headers.get("x-ratelimit-reset-tokens", "3s"))
            with self.lock:
                self.sleep_until = max(self.sleep_until, time.time() + reset + 0.3)

def _parse_reset(s):
    s = (s or "").strip()
    total = 0.0
    for num, unit in re.findall(r"([\d.]+)\s*(ms|m|s|h)", s):
        v = float(num)
        total += {"ms": v / 1000, "s": v, "m": v * 60, "h": v * 3600}[unit]
    return total or 2.0

LIMITER = None

SYS_PROMPT = (
    "You extract medical/dental college fee information from Indian college website "
    "text. Return STRICT JSON only. Prefer MBBS/BDS govt-quota annual fees when present. "
    "Amounts must include the rupee figure exactly as written (e.g. '₹1,25,000' or "
    "'Rs. 25000'). If a value is genuinely not present in the text, use null. Do NOT invent."
)

SCHEMA_HINT = """Return JSON with EXACTLY these keys:
{
  "found": boolean,                // true if any concrete fee figure was present
  "tuition_fee": string|null,      // annual tuition / academic fee (course-level, prefer MBBS/BDS)
  "hostel_fee": string|null,       // hostel / accommodation fee
  "other_fees": string|null,       // any other fees (caution deposit, exam, mess, misc) as one line
  "total_fee": string|null,        // total / grand total per year if stated
  "period": string|null,           // "per year" / "one time" / "per semester" etc.
  "course": string|null,           // course the fee applies to, if stated (e.g. "MBBS")
  "source_url": string|null,       // which of the given [PAGE: url] blocks the fees came from
  "confidence": "high"|"medium"|"low",
  "notes": string|null             // <=160 chars, e.g. quota/category or caveats
}"""

def build_context(pages, max_chars=5000):
    parts, total = [], 0
    for url, txt in pages:
        block = f"[PAGE: {url}]\n{txt}\n"
        if total + len(block) > max_chars:
            block = block[: max_chars - total]
        parts.append(block)
        total += len(block)
        if total >= max_chars:
            break
    return "\n".join(parts)

def llm_extract(college, pages):
    ctx = build_context(pages)
    user = (f"College: {college}\n\n{SCHEMA_HINT}\n\n"
            f"--- SOURCE TEXT (search snippets + pages) ---\n{ctx}")
    payload = {
        "model": EXTRACT_MODEL,
        "messages": [{"role": "system", "content": SYS_PROMPT},
                     {"role": "user", "content": user}],
        "temperature": 0,
        "max_tokens": 400,
        "response_format": {"type": "json_object"},
    }
    hdr = {"Authorization": f"Bearer {EXTRACT_KEY}", "Content-Type": "application/json"}
    last = ""
    for attempt in range(6):
        LIMITER.acquire()
        try:
            r = _rq.post(f"{EXTRACT_BASE}/chat/completions", headers=hdr, json=payload, timeout=60)
        except Exception as e:
            last = f"{type(e).__name__}: {str(e)[:120]}"
            time.sleep(2 * (attempt + 1))
            continue
        LIMITER.note_headers(r.headers)
        if r.status_code == 429:
            wait = _parse_reset(r.headers.get("retry-after", "")) or _parse_reset(
                (re.search(r"try again in ([\d.a-z]+)", r.text) or [None, "5s"])[1])
            # A short wait = per-minute TPM; park and retry. A long wait = the daily
            # token cap — give up fast so the caller can DEFER (resume next run).
            if wait > 90:
                return {"found": False, "error": f"llm_error: 429 daily_cap reset~{int(wait)}s"}
            with LIMITER.lock:
                LIMITER.sleep_until = max(LIMITER.sleep_until, time.time() + wait + 0.5)
            last = "429 rate_limited"
            continue
        if r.status_code >= 400:
            last = f"http_{r.status_code}: {r.text[:140]}"
            if r.status_code in (500, 502, 503):
                time.sleep(2 * (attempt + 1)); continue
            return {"found": False, "error": f"llm_error: {last}"}
        try:
            data = json.loads(r.json()["choices"][0]["message"]["content"])
            return data
        except Exception as e:
            last = f"parse: {str(e)[:120]}"
            time.sleep(1)
            continue
    return {"found": False, "error": f"llm_error: {last or 'exhausted retries'}"}

# ----------------------------------------------------------------------------- pipeline
def process_one(rec):
    num, college, state, city, typ, site = rec
    out = {"num": num, "college": college, "state": state, "city": city,
           "type": typ, "website": site}
    pages, status = gather_pages(site, None)
    out["status"] = status
    reachable = [(u, t) for (u, t) in pages if t and t.strip()]
    joined = "\n".join(t for _u, t in reachable)
    if not reachable:
        out.update({"found": False, "reason": "no_reachable_pages"})
        return out
    if not AMOUNT_RE.search(joined):
        out.update({"found": False, "reason": "no_fee_figures_found",
                    "pages_tried": len(reachable)})
        return out
    data = llm_extract(college, reachable)
    out.update(data)
    out["pages_tried"] = len(reachable)
    return out

def load_targets(path, scope):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Colleges to scrape"]
    targets = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = (list(r) + [None] * 11)[:11]
        num, college, state, city, typ, site, fees, rev, cr, al, what = row
        site = (site or "").strip()
        if not site:
            continue
        fees = (fees or "").strip().upper()
        what = (what or "").strip().lower()
        needs_fees = fees == "MISSING" or "fee" in what
        if scope == "missing" and not needs_fees:
            continue
        targets.append((str(num), (college or "").strip(), (state or "").strip(),
                        (city or "").strip(), (typ or "").strip(), site))
    return targets

def write_outputs(results, out_xlsx, out_csv):
    results = sorted(results, key=lambda d: int(d["num"]) if str(d.get("num", "")).isdigit() else 0)
    cols = ["#", "College", "State", "City", "Type", "Website",
            "Tuition Fee", "Hostel Fee", "Other Fees", "Total Fee", "Period",
            "Course", "Confidence", "Source URL", "Notes", "Status"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fees extracted"
    ws.append(cols)
    def cell(d, k):
        v = d.get(k)
        return "" if v is None else str(v)
    for d in results:
        found = d.get("found")
        status = ("FOUND" if found else "NOT FOUND") + (
            "" if not d.get("reason") else f" ({d['reason']})") + (
            "" if not d.get("error") else f" ({d['error']})")
        ws.append([
            d.get("num", ""), d.get("college", ""), d.get("state", ""), d.get("city", ""),
            d.get("type", ""), d.get("website", ""),
            cell(d, "tuition_fee"), cell(d, "hostel_fee"), cell(d, "other_fees"),
            cell(d, "total_fee"), cell(d, "period"), cell(d, "course"),
            cell(d, "confidence"), cell(d, "source_url"), cell(d, "notes"), status,
        ])
    widths = [5, 42, 16, 14, 12, 34, 22, 20, 30, 18, 12, 10, 11, 34, 34, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(out_xlsx)
    import csv
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for d in results:
            found = d.get("found")
            status = ("FOUND" if found else "NOT FOUND")
            w.writerow([d.get("num", ""), d.get("college", ""), d.get("state", ""),
                        d.get("city", ""), d.get("type", ""), d.get("website", ""),
                        cell(d, "tuition_fee"), cell(d, "hostel_fee"), cell(d, "other_fees"),
                        cell(d, "total_fee"), cell(d, "period"), cell(d, "course"),
                        cell(d, "confidence"), cell(d, "source_url"), cell(d, "notes"), status])

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="/home/swaraj-sah/Downloads/MedCounsel-data-gaps-websites-updated.xlsx")
    ap.add_argument("--outdir", default="/home/swaraj-sah/Downloads")
    ap.add_argument("--scope", choices=["missing", "all"], default="missing")
    ap.add_argument("--limit", type=int, default=0, help="0 = all targets")
    ap.add_argument("--workers", type=int, default=8)
    ap.add_argument("--llm-rpm", type=int, default=25)
    ap.add_argument("--checkpoint", default=None)
    args = ap.parse_args()

    global LIMITER
    LIMITER = RateLimiter(args.llm_rpm)

    if not AI_KEY:
        print("ERROR: no AI_API_KEY in .env", file=sys.stderr)
        sys.exit(1)

    ckpt = args.checkpoint or os.path.join(args.outdir, "fee_extract_checkpoint.jsonl")
    out_xlsx = os.path.join(args.outdir, "MedCounsel-fees-extracted.xlsx")
    out_csv = os.path.join(args.outdir, "MedCounsel-fees-extracted.csv")

    targets = load_targets(args.input, args.scope)
    if args.limit:
        targets = targets[: args.limit]

    done = {}
    if os.path.exists(ckpt):
        for line in open(ckpt, encoding="utf-8"):
            line = line.strip()
            if not line:
                continue
            try:
                d = json.loads(line)
                done[d["num"]] = d
            except Exception:
                pass
    todo = [t for t in targets if t[0] not in done]
    print(f"[fees] targets={len(targets)} already_done={len(done)} todo={len(todo)} "
          f"scope={args.scope} model={AI_MODEL}", flush=True)

    lock = threading.Lock()
    ck = open(ckpt, "a", encoding="utf-8")
    completed = 0
    found_ct = 0
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(process_one, t): t for t in todo}
        for fut in as_completed(futs):
            t = futs[fut]
            try:
                res = fut.result()
            except Exception as e:
                res = {"num": t[0], "college": t[1], "state": t[2], "city": t[3],
                       "type": t[4], "website": t[5], "found": False,
                       "error": f"worker_crash: {type(e).__name__}: {str(e)[:120]}"}
            with lock:
                ck.write(json.dumps(res, ensure_ascii=False) + "\n")
                ck.flush()
                done[res["num"]] = res
                completed += 1
                if res.get("found"):
                    found_ct += 1
                if completed % 10 == 0 or completed == len(todo):
                    rate = completed / max(time.time() - t0, 1)
                    eta = (len(todo) - completed) / max(rate, 1e-6) / 60
                    print(f"[fees] {completed}/{len(todo)} done | found so far "
                          f"{found_ct} | {rate*60:.1f}/min | ETA {eta:.1f} min "
                          f"| last: {res.get('college','')[:36]} -> "
                          f"{'FOUND' if res.get('found') else res.get('reason') or res.get('error') or 'not found'}",
                          flush=True)
    ck.close()

    write_outputs(list(done.values()), out_xlsx, out_csv)
    total_found = sum(1 for d in done.values() if d.get("found"))
    print(f"\n[fees] COMPLETE. processed={len(done)} found_fees={total_found}", flush=True)
    print(f"[fees] Excel: {out_xlsx}", flush=True)
    print(f"[fees] CSV:   {out_csv}", flush=True)
    print(f"[fees] checkpoint: {ckpt}", flush=True)

if __name__ == "__main__":
    main()
